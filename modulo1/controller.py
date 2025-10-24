from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path
import os


class PDFExtractorController:
    """Controlador para el módulo de datos de estudiantes"""
    
    def __init__(self, model, view):
        self.model = model
        self.view = view
        
        # Conectar botones
        self.view.set_button_commands(
            load_cmd=self.load_pdfs,
            extract_cmd=self.extract_data,
            generate_cmd=self.generate_excel,
            debug_cmd=self.debug_pdf,
            clear_cmd=self.clear_data
        )
    
    def load_pdfs(self):
        """Carga archivos PDF"""
        files = filedialog.askopenfilenames(
            title="Seleccionar archivos PDF (Estudiantes)",
            filetypes=[("Archivos PDF", "*.pdf"), ("Todos los archivos", "*.*")]
        )
        
        if files:
            total = self.model.add_pdf_files(files)
            self.view.update_file_list(self.model.get_pdf_names())
            self.view.update_status(f"Cargados {total} archivos PDF")
    
    def extract_data(self):
        """Extrae datos de los PDFs"""
        if not self.model.has_files():
            messagebox.showwarning("Advertencia", "No hay archivos PDF cargados")
            return
        
        # Mostrar loading
        self.view.show_loading("Extrayendo datos de estudiantes...")
        
        try:
            self.view.clear_data_tree()
            self.view.update_status("Extrayendo datos...")
            self.view.root.update()
            
            extracted_data = self.model.process_all_pdfs()
            
            # PRIMERO ocultar loading
            self.view.hide_loading()
            
            # LUEGO mostrar datos
            for i, data in enumerate(extracted_data, 1):
                if 'error' in data:
                    self.view.add_data_to_tree(i, f"ERROR: {data['archivo']}", "", "")
                else:
                    self.view.add_data_to_tree(
                        i, 
                        data['nombres'], 
                        data['dni'], 
                        data['nivel_riesgo']
                    )
            
            self.view.update_status(f"Datos extraídos de {len(extracted_data)} archivos")
            
            if extracted_data:
                messagebox.showinfo(
                    "Éxito", 
                    f"Se extrajeron datos de {len(extracted_data)} archivos.\n"
                    "Revise los datos en la vista previa."
                )
        
        except Exception as e:
            # Asegurar que se oculte incluso con error
            self.view.hide_loading()
            messagebox.showerror("Error", f"Error durante la extracción: {str(e)}")
            self.view.update_status("Error en la extracción")
    
    def generate_excel(self):
        """Genera archivo Excel"""
        if not self.model.has_data():
            messagebox.showwarning(
                "Advertencia",
                "No hay datos extraídos. Primero extraiga los datos."
            )
            return
        
        # Solicitar ubicación con ruta por defecto en Escritorio
        desktop_path = Path.home() / "Desktop" / "datos_estudiantes.xlsx"
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="datos_estudiantes.xlsx",
            initialdir=str(Path.home() / "Desktop")
        )
        
        if not file_path:
            return
        
        # Mostrar loading
        self.view.show_loading("Generando archivo Excel...")
        
        try:
            final_file_path = self._create_excel_file(file_path)
            
            # PRIMERO ocultar loading
            self.view.hide_loading()
            
            # LUEGO mostrar mensajes
            self.view.update_status(f"Excel generado: {Path(final_file_path).name}")
            messagebox.showinfo(
                "Éxito",
                f"Archivo Excel generado correctamente:\n{final_file_path}"
            )
            
            # Preguntar si desea abrir
            if messagebox.askyesno("Abrir archivo", "¿Desea abrir el archivo Excel?"):
                try:
                    os.startfile(final_file_path)
                except Exception as e:
                    messagebox.showwarning(
                        "Aviso", 
                        f"No se pudo abrir el archivo automáticamente:\n{str(e)}\n\n"
                        f"Puede abrirlo manualmente desde:\n{final_file_path}"
                    )
        
        except Exception as e:
            self.view.hide_loading()
            error_msg = f"Error al generar Excel: {str(e)}"
            
            if "permiso" in str(e).lower() or "permission" in str(e).lower():
                error_msg += "\n\nSugerencias:\n"
                error_msg += "• Cierre Excel si tiene abierto un archivo con el mismo nombre\n"
                error_msg += "• Intente guardar en una carpeta diferente\n"
                error_msg += "• Verifique los permisos de la carpeta destino"
            
            messagebox.showerror("Error", error_msg)
    
    def _create_excel_file(self, file_path):
        """Crea el archivo Excel con los datos - Versión robusta"""
        try:
            file_path = Path(file_path)
            parent_dir = file_path.parent
            
            # Verificar que el directorio existe y tiene permisos de escritura
            if not parent_dir.exists():
                parent_dir.mkdir(parents=True, exist_ok=True)
            
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Datos Estudiantes"
            
            # Estilos
            header_fill = PatternFill(
                start_color="366092",
                end_color="366092",
                fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            # Encabezados
            headers = ["N°", "Nombres y Apellidos", "DNI", "Nivel de Riesgo Social"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Datos
            extracted_data = self.model.get_extracted_data()
            for i, data in enumerate(extracted_data, 2):
                ws.cell(row=i, column=1, value=i-1)
                ws.cell(row=i, column=2, value=data['nombres'])
                ws.cell(row=i, column=3, value=data['dni'])
                ws.cell(row=i, column=4, value=data['nivel_riesgo'])
            
            # Ajustar anchos
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 25
            
            # Guardar
            wb.save(str(file_path))
            return str(file_path)
            
        except Exception as e:
            raise Exception(f"Error al guardar Excel: {str(e)}")
    
    def clear_data(self):
        """Limpia todos los datos"""
        self.model.clear_files()
        self.view.update_file_list([])
        self.view.clear_data_tree()
        self.view.update_status("Datos limpiados")
    
    def debug_pdf(self):
        """Muestra el texto del primer PDF para debug"""
        if not self.model.has_files():
            messagebox.showwarning("Advertencia", "No hay archivos PDF cargados")
            return
        
        # Mostrar loading
        self.view.show_loading("Extrayendo texto del PDF...")
        
        try:
            pdf_files = self.model.get_pdf_files()
            pdf_path = pdf_files[0]
            text = self.model.extract_text_from_pdf(pdf_path)
            
            # PRIMERO ocultar loading
            self.view.hide_loading()
            
            # LUEGO mostrar ventana de debug
            pdf_name = Path(pdf_path).name
            self.view.create_debug_window(pdf_name, text)
        
        except Exception as e:
            self.view.hide_loading()
            messagebox.showerror("Error", f"Error en debug: {str(e)}")