from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path
import os


class TransSegenController:
    """Controlador para el módulo Trans-Segen"""
    
    def __init__(self, model, view):
        self.model = model
        self.view = view
        
        # Conectar botones
        self.view.set_button_commands(
            load_cmd=self.load_pdfs,
            extract_cmd=self.extract_data,
            generate_cmd=self.generate_excel,
            debug_cmd=self.debug_pdf,
            clear_cmd=self.clear_data
        )
    
    def load_pdfs(self):
        """Carga archivos PDF"""
        files = filedialog.askopenfilenames(
            title="Seleccionar archivos PDF (Trans-Segen)",
            filetypes=[("Archivos PDF", "*.pdf"), ("Todos los archivos", "*.*")]
        )
        
        if files:
            total = self.model.add_pdf_files(files)
            self.view.update_file_list(self.model.get_pdf_names())
            self.view.update_status(f"Cargados {total} archivos PDF")
    
    def extract_data(self):
        """Extrae datos usando OCR"""
        if not self.model.has_files():
            messagebox.showwarning("Advertencia", "No hay archivos PDF cargados")
            return
        
        # Advertir sobre el tiempo de procesamiento
        response = messagebox.askyesno(
            "Procesamiento OCR",
            "El procesamiento con OCR puede tardar varios minutos.\n"
            "¿Desea continuar?",
            icon='question'
        )
        
        if not response:
            return
        
        # Mostrar loading
        self.view.show_loading("Procesando con OCR... Esto puede tardar varios minutos")
        
        try:
            self.view.clear_data_tree()
            self.view.update_status("Procesando con OCR... Por favor espere...")
            self.view.root.update()
            
            extracted_data = self.model.process_all_pdfs()
            
            # **PRIMERO ocultar el loading**
            self.view.hide_loading()
            
            # **DESPUÉS mostrar los datos en el treeview**
            for i, data in enumerate(extracted_data, 1):
                if 'error' in data:
                    self.view.add_data_to_tree(
                        i,
                        f"ERROR: {data['archivo']}",
                        ""
                    )
                else:
                    self.view.add_data_to_tree(
                        i,
                        data['nombres'],
                        data['nro_transegen']
                    )
            
            self.view.update_status(f"Datos extraídos de {len(extracted_data)} archivos")
            
            if extracted_data:
                # Contar errores
                errores = sum(1 for d in extracted_data if 'error' in d)
                exitosos = len(extracted_data) - errores
                
                msg = f"Procesamiento completado:\n"
                msg += f"✓ Exitosos: {exitosos}\n"
                if errores > 0:
                    msg += f"✗ Con errores: {errores}\n"
                msg += "\nRevise los datos en la vista previa."
                
                # **AHORA SÍ mostrar el messagebox (sin loading bloqueando)**
                messagebox.showinfo("Procesamiento Completado", msg)
        
        except Exception as e:
            # **Ocultar loading ANTES de mostrar el error**
            self.view.hide_loading()
            messagebox.showerror("Error", f"Error durante la extracción: {str(e)}")
            self.view.update_status("Error en la extracción")
        
    def generate_excel(self):
        """Genera archivo Excel"""
        if not self.model.has_data():
            messagebox.showwarning(
                "Advertencia",
                "No hay datos extraídos. Primero extraiga los datos."
            )
            return
        
        # Solicitar ubicación
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="datos_transegen.xlsx"
        )
        
        if not file_path:
            return
        
        try:
            self._create_excel_file(file_path)
            
            self.view.update_status(f"Excel generado: {Path(file_path).name}")
            messagebox.showinfo(
                "Éxito",
                f"Archivo Excel generado correctamente:\n{file_path}"
            )
            
            # Preguntar si desea abrir
            if messagebox.askyesno("Abrir archivo", "¿Desea abrir el archivo Excel?"):
                os.startfile(file_path)
        
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar Excel: {str(e)}")
    
    def _create_excel_file(self, file_path):
        """Crea el archivo Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos Trans-Segen"
        
        # Estilos
        header_fill = PatternFill(
            start_color="2E7D32",
            end_color="2E7D32",
            fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Encabezados
        headers = ["N°", "Nombres y Apellidos", "Nro Trans-Segen"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Datos
        extracted_data = self.model.get_extracted_data()
        for i, data in enumerate(extracted_data, 2):
            ws.cell(row=i, column=1, value=i-1)
            ws.cell(row=i, column=2, value=data['nombres'])
            ws.cell(row=i, column=3, value=data['nro_transegen'])
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 35
        
        # Guardar
        wb.save(file_path)
    
    def clear_data(self):
        """Limpia todos los datos"""
        self.model.clear_files()
        self.view.update_file_list([])
        self.view.clear_data_tree()
        self.view.update_status("Datos limpiados")
    
    def debug_pdf(self):
        """Muestra el texto OCR del primer PDF"""
        if not self.model.has_files():
            messagebox.showwarning("Advertencia", "No hay archivos PDF cargados")
            return
        
        # Mostrar loading para debug
        self.view.show_loading("Extrayendo texto con OCR...")
        
        try:
            self.view.update_status("Procesando OCR para debug...")
            
            pdf_files = self.model.get_pdf_files()
            pdf_path = pdf_files[0]
            text = self.model.extract_text_from_pdf_ocr(pdf_path)
            
            # **PRIMERO ocultar loading**
            self.view.hide_loading()
            
            # **DESPUÉS mostrar la ventana de debug**
            pdf_name = Path(pdf_path).name
            self.view.create_debug_window(pdf_name, text)
            self.view.update_status("Debug completado")
        
        except Exception as e:
            # **Ocultar loading ANTES de mostrar error**
            self.view.hide_loading()
            messagebox.showerror("Error", f"Error en debug: {str(e)}")
            self.view.update_status("Error en debug")